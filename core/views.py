import io
import json

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, Http404
from django.views.decorators.http import require_POST
from django.db.models import Count

from .forms import RegisterForm, LoginForm
from .models import (
    AnalysisSession, RequirementGap, TestScenario,
    FeedbackItem, TestCondition, DetailedTestCase,
    Workspace, WorkspaceMembership, WorkspaceDraftInput,
)
from .services.qa_analyzer import QAAnalyzer
from .services.detailed_case_generator import DetailedCaseGenerator
from .services.training_data_manager import TrainingDataManager
from .services.gemini_service import GeminiError


# ─────────────────────────────────────────────
# Public pages
# ─────────────────────────────────────────────

def home_view(request):
    return render(request, 'home.html')


# ─────────────────────────────────────────────
# Authentication
# ─────────────────────────────────────────────

def register_view(request):
    if request.user.is_authenticated:
        return redirect('workspace')

    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            messages.success(request, f'Registration successful. Welcome, {user.username}!')
            return redirect('workspace')
    else:
        form = RegisterForm()

    return render(request, 'register.html', {'form': form})


def login_view(request):
    if request.user.is_authenticated:
        return redirect('workspace')

    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('workspace')
    else:
        form = LoginForm()

    return render(request, 'login.html', {'form': form})


def logout_view(request):
    logout(request)
    return redirect('home')


# ─────────────────────────────────────────────
# Main workspace
# ─────────────────────────────────────────────

def _my_workspaces(user):
    """Team workspaces this user belongs to, newest first, with member counts."""
    ws_ids = WorkspaceMembership.objects.filter(user=user).values_list('workspace_id', flat=True)
    return Workspace.objects.filter(id__in=ws_ids).annotate(
        member_count=Count('memberships', distinct=True)
    ).order_by('-created_at')


@login_required
def workspace_view(request):
    history = AnalysisSession.objects.filter(
        user=request.user, workspace__isnull=True
    ).order_by('-is_pinned', '-created_at')
    return render(request, 'workspace.html', {
        'history': history,
        'my_workspaces': _my_workspaces(request.user),
        'can_delete_chats': True,  # personal chats are always the user's own
    })


# ─────────────────────────────────────────────
# Shared workspace (multi-user)
# ─────────────────────────────────────────────

@login_required
def shared_workspace_view(request, workspace_id):
    workspace = get_object_or_404(Workspace, workspace_id=workspace_id)

    membership, just_joined = WorkspaceMembership.objects.get_or_create(
        workspace=workspace,
        user=request.user,
        defaults={'role': 'member'},
    )

    history = AnalysisSession.objects.filter(
        workspace=workspace
    ).order_by('-is_pinned', '-created_at')

    members = WorkspaceMembership.objects.filter(
        workspace=workspace
    ).select_related('user')

    return render(request, 'workspace.html', {
        'history': history,
        'workspace': workspace,
        'members': members,
        'just_joined': just_joined,
        'my_workspaces': _my_workspaces(request.user),
        'can_delete_chats': workspace.owner == request.user,
        'is_owner': workspace.owner == request.user,
    })


# ─────────────────────────────────────────────
# Create workspace
# ─────────────────────────────────────────────

@login_required
@require_POST
def create_workspace_view(request):
    try:
        body = json.loads(request.body)
        name = body.get('name', '').strip() or 'Team Workspace'
    except (json.JSONDecodeError, AttributeError):
        name = 'Team Workspace'

    workspace = Workspace.objects.create(name=name[:200], owner=request.user)
    WorkspaceMembership.objects.create(workspace=workspace, user=request.user, role='owner')

    return JsonResponse({'workspace_id': workspace.workspace_id, 'name': workspace.name})


# ─────────────────────────────────────────────
# Rename a workspace (any member)
# ─────────────────────────────────────────────

@login_required
@require_POST
def rename_workspace_view(request):
    try:
        body = json.loads(request.body)
        workspace_id = body.get('workspace_id', '').strip()
        name = body.get('name', '').strip()
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid request body.'}, status=400)

    if not workspace_id or not name:
        return JsonResponse({'error': 'workspace_id and name are required.'}, status=400)

    workspace = get_object_or_404(Workspace, workspace_id=workspace_id)
    if not WorkspaceMembership.objects.filter(workspace=workspace, user=request.user).exists():
        return JsonResponse({'error': 'Only a workspace member can rename it.'}, status=403)

    workspace.name = name[:200]
    workspace.save(update_fields=['name'])
    return JsonResponse({'status': 'ok', 'name': workspace.name})


# ─────────────────────────────────────────────
# Leave / quit a workspace (any member)
# ─────────────────────────────────────────────

@login_required
@require_POST
def leave_workspace_view(request):
    try:
        body = json.loads(request.body)
        workspace_id = body.get('workspace_id', '').strip()
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid request body.'}, status=400)

    if not workspace_id:
        return JsonResponse({'error': 'workspace_id is required.'}, status=400)

    workspace = get_object_or_404(Workspace, workspace_id=workspace_id)
    if workspace.owner == request.user:
        return JsonResponse(
            {'error': 'The owner cannot leave. Use Delete Workspace instead.'}, status=400
        )

    # Remove the member and clean up the data they own in this workspace
    # (their draft contribution). Shared team chats stay for the rest of the team.
    WorkspaceMembership.objects.filter(workspace=workspace, user=request.user).delete()
    WorkspaceDraftInput.objects.filter(workspace=workspace, user=request.user).delete()
    return JsonResponse({'status': 'left'})


# ─────────────────────────────────────────────
# Add a member by username (any member)
# ─────────────────────────────────────────────

@login_required
@require_POST
def add_member_view(request):
    try:
        body = json.loads(request.body)
        workspace_id = body.get('workspace_id', '').strip()
        username = body.get('username', '').strip()
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid request body.'}, status=400)

    if not workspace_id or not username:
        return JsonResponse({'error': 'workspace_id and username are required.'}, status=400)

    workspace = get_object_or_404(Workspace, workspace_id=workspace_id)
    if not WorkspaceMembership.objects.filter(workspace=workspace, user=request.user).exists():
        return JsonResponse({'error': 'Only a workspace member can add members.'}, status=403)

    try:
        user = User.objects.get(username__iexact=username)
    except User.DoesNotExist:
        return JsonResponse({'error': 'No user named "%s".' % username}, status=404)

    _, created = WorkspaceMembership.objects.get_or_create(
        workspace=workspace, user=user, defaults={'role': 'member'}
    )
    if not created:
        return JsonResponse({'error': '%s is already a member.' % user.username}, status=400)

    return JsonResponse({
        'status': 'added',
        'username': user.username,
        'member_count': WorkspaceMembership.objects.filter(workspace=workspace).count(),
    })


# ─────────────────────────────────────────────
# Remove a member (any member; the owner cannot be removed)
# ─────────────────────────────────────────────

@login_required
@require_POST
def remove_member_view(request):
    try:
        body = json.loads(request.body)
        workspace_id = body.get('workspace_id', '').strip()
        username = body.get('username', '').strip()
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid request body.'}, status=400)

    if not workspace_id or not username:
        return JsonResponse({'error': 'workspace_id and username are required.'}, status=400)

    workspace = get_object_or_404(Workspace, workspace_id=workspace_id)
    if not WorkspaceMembership.objects.filter(workspace=workspace, user=request.user).exists():
        return JsonResponse({'error': 'Only a workspace member can remove members.'}, status=403)
    if username.lower() == workspace.owner.username.lower():
        return JsonResponse({'error': 'The owner cannot be removed.'}, status=400)

    deleted, _ = WorkspaceMembership.objects.filter(
        workspace=workspace, user__username__iexact=username
    ).delete()
    if not deleted:
        return JsonResponse({'error': '%s is not a member.' % username}, status=404)

    return JsonResponse({
        'status': 'removed',
        'username': username,
        'member_count': WorkspaceMembership.objects.filter(workspace=workspace).count(),
    })


# ─────────────────────────────────────────────
# Delete a whole workspace + all its chats (any member)
# ─────────────────────────────────────────────

@login_required
@require_POST
def delete_workspace_view(request):
    try:
        body = json.loads(request.body)
        workspace_id = body.get('workspace_id', '').strip()
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid request body.'}, status=400)

    if not workspace_id:
        return JsonResponse({'error': 'workspace_id is required.'}, status=400)

    workspace = get_object_or_404(Workspace, workspace_id=workspace_id)
    if not WorkspaceMembership.objects.filter(workspace=workspace, user=request.user).exists():
        return JsonResponse({'error': 'Only a workspace member can delete the workspace.'}, status=403)

    # Sessions use SET_NULL, so remove them explicitly for a complete deletion;
    # this cascades to their scenarios/conditions/gaps/feedback. Memberships
    # cascade automatically when the workspace is deleted.
    AnalysisSession.objects.filter(workspace=workspace).delete()
    workspace.delete()

    return JsonResponse({'status': 'deleted'})


# ─────────────────────────────────────────────
# Team input draft — members contribute, owner generates
# ─────────────────────────────────────────────

@login_required
def workspace_draft_view(request, workspace_id):
    """List every member's draft contribution, for live polling."""
    workspace = get_object_or_404(Workspace, workspace_id=workspace_id)
    if not WorkspaceMembership.objects.filter(workspace=workspace, user=request.user).exists():
        raise Http404()

    inputs = WorkspaceDraftInput.objects.filter(workspace=workspace).select_related('user')
    return JsonResponse({
        'inputs': [
            {'username': d.user.username, 'text': d.text, 'is_me': d.user_id == request.user.id}
            for d in inputs
        ],
        'is_owner': workspace.owner_id == request.user.id,
    })


@login_required
@require_POST
def save_draft_view(request, workspace_id):
    """Save a draft contribution. Any member may edit their own or another
    member's contribution (pass `username` to target someone else)."""
    workspace = get_object_or_404(Workspace, workspace_id=workspace_id)
    if not WorkspaceMembership.objects.filter(workspace=workspace, user=request.user).exists():
        raise Http404()
    try:
        body = json.loads(request.body)
        text = body.get('text', '')
        username = body.get('username', '').strip()
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid request.'}, status=400)

    target = request.user
    if username and username.lower() != request.user.username.lower():
        target = User.objects.filter(username__iexact=username).first()
        if target is None or not WorkspaceMembership.objects.filter(
            workspace=workspace, user=target
        ).exists():
            return JsonResponse({'error': 'That member is not in this workspace.'}, status=404)

    WorkspaceDraftInput.objects.update_or_create(
        workspace=workspace, user=target, defaults={'text': text}
    )
    return JsonResponse({'status': 'ok'})


@login_required
@require_POST
def generate_from_draft_view(request, workspace_id):
    """Any member: combine all members' drafts and run the QA analysis."""
    workspace = get_object_or_404(Workspace, workspace_id=workspace_id)
    if not WorkspaceMembership.objects.filter(workspace=workspace, user=request.user).exists():
        raise Http404()

    inputs = WorkspaceDraftInput.objects.filter(workspace=workspace).order_by('updated_at')
    combined = '\n\n'.join(d.text.strip() for d in inputs if d.text.strip())
    if not combined:
        return JsonResponse({'error': 'No team input to generate from.'}, status=400)

    try:
        qa_result = QAAnalyzer().analyze(combined)
    except GeminiError as e:
        return JsonResponse({'error': str(e)}, status=e.status)

    session = TrainingDataManager().save_qa_session(
        user=request.user, requirements_text=combined, qa_result=qa_result
    )
    session.workspace = workspace
    session.save(update_fields=['workspace'])

    # Drafts are kept (not cleared) so the team can refine and regenerate.
    requires = session.accuracy_score < 80
    suggested = qa_result.get('suggested_requirement', '')
    return JsonResponse(_build_session_response(
        session,
        requires_decision=requires,
        suggested=suggested if requires else '',
    ))


# ─────────────────────────────────────────────
# Serialization helpers
# ─────────────────────────────────────────────

def _serialize_scenarios(session):
    return [
        {
            'db_id': s.id,
            'id': s.scenario_id,
            'requirement_ref': s.requirement_ref,
            'condition_ref': s.condition_ref,
            'description': s.description,
            'preconditions': s.preconditions,
            'steps': s.get_steps(),
            'expected_result': s.expected_result,
            'type': s.scenario_type,
            'priority': s.priority,
            'is_done': s.is_done,
        }
        for s in session.test_scenarios.all()
    ]


def _serialize_conditions(session):
    return [
        {
            'db_id': c.id,
            'condition_id': c.condition_id,
            'requirement_ref': c.requirement_ref,
            'description': c.description,
            'type': c.condition_type,
            'priority': c.priority,
        }
        for c in session.test_conditions.all()
    ]


def _session_requirements(session):
    """Return the list of requirements for a session, handling both the new
    multi-requirement format and the old single-requirement format."""
    info = session.get_extracted_info() or {}
    if isinstance(info, dict) and isinstance(info.get('requirements'), list) and info['requirements']:
        return info['requirements']
    if isinstance(info, dict) and (info.get('requirement_id') or info.get('actors') is not None):
        return [info]  # old format: a single requirement dict
    return [{'requirement_id': session.requirement_id or 'REQ-001'}]


def _serialize_gaps(session):
    return [
        {
            'db_id': g.id,
            'issue_id': g.issue_id,
            'issue_type': g.issue_type,
            'description': g.description,
            'suggested_clarification': g.suggested_clarification,
        }
        for g in session.gaps.all()
    ]


def _serialize_quality_assessment(session):
    return {
        'clarity_score': session.clarity_score,
        'completeness_score': session.completeness_score,
        'testability_score': session.testability_score,
        'overall_score': session.accuracy_score,
        'severity': session.severity or 'Medium',
        'positive_aspects': [
            {'db_id': f.id, 'text': f.message}
            for f in session.feedback_items.filter(feedback_type='positive')
        ],
        'warnings': [
            {'db_id': f.id, 'text': f.message}
            for f in session.feedback_items.filter(feedback_type='warning')
        ],
    }


def _serialize_detailed_cases(session):
    """Serialize any saved detailed test cases (Section 6), with per-step done state."""
    cases = []
    for s in session.test_scenarios.all():
        try:
            dc = s.detailed_case
        except DetailedTestCase.DoesNotExist:
            continue
        steps = dc.get_steps()
        done = dc.get_steps_done()
        done = (done + [False] * len(steps))[:len(steps)]  # align to step count
        cases.append({
            'db_id': dc.id,
            'scenario_id': s.scenario_id,
            'test_data': dc.test_data,
            'steps': steps,
            'steps_done': done,
            'expected_results': dc.expected_results,
            'postconditions': dc.postconditions,
        })
    return cases


def _build_session_response(session, requires_decision=False, suggested=''):
    """Build the standard JSON response payload for a session."""
    return {
        'session_id': session.id,
        'requirements_text': session.requirements_text,
        'score': session.accuracy_score,
        'score_label': session.get_score_label(),
        'score_color': session.get_score_color(),
        'requirements': _session_requirements(session),
        'quality_assessment': _serialize_quality_assessment(session),
        'test_conditions': _serialize_conditions(session),
        'gaps': _serialize_gaps(session),
        'scenarios': _serialize_scenarios(session),
        'detailed_cases': _serialize_detailed_cases(session),
        'requires_user_decision': requires_decision,
        'suggested_requirement': suggested,
        'team_notes': session.team_notes,
    }


def _get_accessible_session(session_id, user):
    """Return session if user is the owner or a workspace member, else raise Http404."""
    session = get_object_or_404(AnalysisSession, id=session_id)
    if session.user == user:
        return session
    if session.workspace and WorkspaceMembership.objects.filter(
        workspace=session.workspace, user=user
    ).exists():
        return session
    raise Http404()


# ─────────────────────────────────────────────
# Load session
# ─────────────────────────────────────────────

@login_required
def load_session_view(request, session_id):
    session = _get_accessible_session(session_id, request.user)
    return JsonResponse(_build_session_response(session))


# ─────────────────────────────────────────────
# Live workspace state (lightweight polling endpoint)
# ─────────────────────────────────────────────

@login_required
def workspace_state_view(request, workspace_id):
    """
    Cheap JSON snapshot of a shared workspace for near-real-time polling:
    the session list (for the sidebar) and the current member list.
    Makes no LLM calls. Only accessible to members of the workspace.
    """
    workspace = get_object_or_404(Workspace, workspace_id=workspace_id)
    if not WorkspaceMembership.objects.filter(
        workspace=workspace, user=request.user
    ).exists():
        raise Http404()

    sessions = AnalysisSession.objects.filter(
        workspace=workspace
    ).order_by('-is_pinned', '-created_at')

    members = list(
        WorkspaceMembership.objects.filter(workspace=workspace)
        .select_related('user')
        .values_list('user__username', flat=True)
    )

    return JsonResponse({
        'sessions': [
            {
                'id': s.id,
                'title': s.title,
                'score': s.accuracy_score,
                'color': s.get_score_color(),
                'pinned': s.is_pinned,
            }
            for s in sessions
        ],
        'members': members,
        'member_count': len(members),
        'owner': workspace.owner.username,
    })


@login_required
def my_workspaces_state_view(request):
    """Cheap JSON snapshot of the user's team workspaces for live sidebar refresh."""
    return JsonResponse({'workspaces': [
        {'workspace_id': w.workspace_id, 'name': w.name, 'member_count': w.member_count}
        for w in _my_workspaces(request.user)
    ]})


# ─────────────────────────────────────────────
# Analyze requirements — Phase 3 QA redesign
# ─────────────────────────────────────────────

@login_required
@require_POST
def analyze_view(request):
    try:
        body = json.loads(request.body)
        requirements_text = body.get('requirements_text', '').strip()
        workspace_id = body.get('workspace_id', '').strip()
        keep_session_id = body.get('keep_session_id')
    except (json.JSONDecodeError, AttributeError):
        requirements_text = request.POST.get('requirements_text', '').strip()
        workspace_id = ''
        keep_session_id = None

    # User chose "use my original input" on a weak requirement: the session
    # was already scored in stage 1, so skip straight to stage 2 (scenarios)
    # instead of re-extracting the same text again.
    if keep_session_id:
        session = _get_accessible_session(keep_session_id, request.user)
        requirements = session.get_extracted_info().get('requirements', [])
        try:
            qa_result = QAAnalyzer().generate_scenarios_only(requirements)
        except GeminiError as e:
            return JsonResponse({'error': str(e)}, status=e.status)
        TrainingDataManager().reanalyze_session(session, qa_result)
        return JsonResponse(_build_session_response(session))

    if not requirements_text:
        return JsonResponse({'error': 'Requirements text is required.'}, status=400)

    # Resolve workspace (if any)
    workspace = None
    if workspace_id:
        try:
            ws = Workspace.objects.get(workspace_id=workspace_id)
            if WorkspaceMembership.objects.filter(workspace=ws, user=request.user).exists():
                workspace = ws
        except Workspace.DoesNotExist:
            pass

    EDITED_PREFIX = 'EDITED:'
    if requirements_text.startswith(EDITED_PREFIX):
        edited_text = requirements_text[len(EDITED_PREFIX):].strip()
        try:
            # The user already decided — always generate the full report.
            qa_result = QAAnalyzer().analyze(edited_text, force_full=True)
        except GeminiError as e:
            return JsonResponse({'error': str(e)}, status=e.status)
        session = TrainingDataManager().save_qa_session(
            user=request.user,
            requirements_text=edited_text,
            qa_result=qa_result,
        )
        if workspace:
            session.workspace = workspace
            session.save(update_fields=['workspace'])
        resp = _build_session_response(session)
        resp['system_action'] = 'analysis_and_generation'
        return JsonResponse(resp)

    try:
        qa_result = QAAnalyzer().analyze(requirements_text)
    except GeminiError as e:
        return JsonResponse({'error': str(e)}, status=e.status)

    session = TrainingDataManager().save_qa_session(
        user=request.user,
        requirements_text=requirements_text,
        qa_result=qa_result,
    )
    if workspace:
        session.workspace = workspace
        session.save(update_fields=['workspace'])

    overall_score = session.accuracy_score
    suggested = qa_result.get('suggested_requirement', '')
    requires_improvement = overall_score < 80

    return JsonResponse(_build_session_response(
        session,
        requires_decision=requires_improvement,
        suggested=suggested if requires_improvement else '',
    ))


# ─────────────────────────────────────────────
# Section 6: Generate detailed test cases on demand
# ─────────────────────────────────────────────

@login_required
@require_POST
def generate_detailed_cases_view(request):
    try:
        body = json.loads(request.body)
        session_id = body.get('session_id')
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid request body.'}, status=400)

    if not session_id:
        return JsonResponse({'error': 'session_id is required.'}, status=400)

    session = _get_accessible_session(session_id, request.user)

    scenarios = [
        {
            'scenario_id': s.scenario_id,
            'description': s.description,
            'preconditions': s.preconditions,
            'expected_result': s.expected_result,
        }
        for s in session.test_scenarios.all()
    ]

    if not scenarios:
        return JsonResponse({'error': 'No scenarios found for this session.'}, status=400)

    try:
        detailed_cases = DetailedCaseGenerator().generate(scenarios)
    except GeminiError as e:
        return JsonResponse({'error': str(e)}, status=e.status)
    TrainingDataManager().save_detailed_cases(session_id, detailed_cases)

    # Return the saved version (with db ids + per-step done state) so the
    # frontend can render the per-step checkboxes.
    return JsonResponse({
        'session_id': session_id,
        'detailed_cases': _serialize_detailed_cases(session),
    })


# ─────────────────────────────────────────────
# Toggle a single detailed-step's "done" checkbox (saved & shared)
# ─────────────────────────────────────────────

@login_required
@require_POST
def toggle_step_done_view(request):
    try:
        body = json.loads(request.body)
        case_id = body.get('case_id')
        step_index = body.get('step_index')
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid request.'}, status=400)

    if case_id is None or step_index is None:
        return JsonResponse({'error': 'case_id and step_index are required.'}, status=400)

    case = get_object_or_404(DetailedTestCase, id=case_id)
    session = case.scenario.session
    if session.user != request.user and not (
        session.workspace and WorkspaceMembership.objects.filter(
            workspace=session.workspace, user=request.user
        ).exists()
    ):
        return JsonResponse({'error': 'Access denied.'}, status=403)

    steps = case.get_steps()
    done = case.get_steps_done()
    done = (done + [False] * len(steps))[:len(steps)]

    try:
        idx = int(step_index)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid step_index.'}, status=400)
    if idx < 0 or idx >= len(steps):
        return JsonResponse({'error': 'step_index out of range.'}, status=400)

    done[idx] = not done[idx]
    case.steps_done = json.dumps(done)
    case.save(update_fields=['steps_done'])

    return JsonResponse({
        'status': 'ok',
        'case_id': case.id,
        'steps_done': done,
        'done_count': sum(1 for x in done if x),
        'total': len(steps),
    })


# ─────────────────────────────────────────────
# Update a single output field (feedback / gap / condition / scenario)
# ─────────────────────────────────────────────

@login_required
@require_POST
def update_output_field_view(request):
    try:
        body = json.loads(request.body)
        model_name = body.get('model')
        db_id = body.get('db_id')
        field = body.get('field')
        value = body.get('value', '').strip()
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid request.'}, status=400)

    if not all([model_name, db_id, field]):
        return JsonResponse({'error': 'model, db_id, field are required.'}, status=400)

    def _has_access(session):
        if session.user == request.user:
            return True
        return session.workspace and WorkspaceMembership.objects.filter(
            workspace=session.workspace, user=request.user
        ).exists()

    if model_name == 'scenario':
        obj = get_object_or_404(TestScenario, id=db_id)
        if not _has_access(obj.session):
            return JsonResponse({'error': 'Access denied.'}, status=403)
        allowed = ('description', 'preconditions', 'expected_result')
        if field not in allowed:
            return JsonResponse({'error': 'Invalid field.'}, status=400)
        setattr(obj, field, value)
        obj.save(update_fields=[field])

    elif model_name == 'feedback':
        obj = get_object_or_404(FeedbackItem, id=db_id)
        if not _has_access(obj.session):
            return JsonResponse({'error': 'Access denied.'}, status=403)
        if field != 'message':
            return JsonResponse({'error': 'Invalid field.'}, status=400)
        obj.message = value
        obj.save(update_fields=['message'])

    elif model_name == 'gap':
        obj = get_object_or_404(RequirementGap, id=db_id)
        if not _has_access(obj.session):
            return JsonResponse({'error': 'Access denied.'}, status=403)
        allowed = ('description', 'suggested_clarification')
        if field not in allowed:
            return JsonResponse({'error': 'Invalid field.'}, status=400)
        setattr(obj, field, value)
        obj.save(update_fields=[field])

    elif model_name == 'condition':
        obj = get_object_or_404(TestCondition, id=db_id)
        if not _has_access(obj.session):
            return JsonResponse({'error': 'Access denied.'}, status=403)
        if field != 'description':
            return JsonResponse({'error': 'Invalid field.'}, status=400)
        obj.description = value
        obj.save(update_fields=['description'])

    else:
        return JsonResponse({'error': 'Invalid model.'}, status=400)

    return JsonResponse({'status': 'ok'})


# ─────────────────────────────────────────────
# Update team notes on a session
# ─────────────────────────────────────────────

@login_required
@require_POST
def update_team_notes_view(request):
    try:
        body = json.loads(request.body)
        session_id = body.get('session_id')
        notes = body.get('notes', '')
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid request.'}, status=400)

    if not session_id:
        return JsonResponse({'error': 'session_id is required.'}, status=400)

    session = _get_accessible_session(session_id, request.user)
    session.team_notes = notes
    session.save(update_fields=['team_notes'])
    return JsonResponse({'status': 'ok'})


@login_required
def session_notes_view(request, session_id):
    """Lightweight read of a session's team notes, for live polling."""
    session = _get_accessible_session(session_id, request.user)
    return JsonResponse({'team_notes': session.team_notes})


# ─────────────────────────────────────────────
# Delete all history
# ─────────────────────────────────────────────

@login_required
@require_POST
def delete_history_view(request):
    try:
        body = json.loads(request.body)
        workspace_id = body.get('workspace_id', '').strip()
    except (json.JSONDecodeError, AttributeError):
        workspace_id = ''

    if workspace_id:
        workspace = get_object_or_404(Workspace, workspace_id=workspace_id)
        if workspace.owner != request.user:
            return JsonResponse({'error': 'Only the workspace owner can clear history.'}, status=403)
        AnalysisSession.objects.filter(workspace=workspace).delete()
    else:
        AnalysisSession.objects.filter(user=request.user, workspace__isnull=True).delete()

    return JsonResponse({'status': 'cleared', 'message': 'History cleared successfully.'})


# ─────────────────────────────────────────────
# Toggle pin
# ─────────────────────────────────────────────

@login_required
@require_POST
def toggle_pin_view(request):
    try:
        body = json.loads(request.body)
        session_id = body.get('session_id')
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid request body.'}, status=400)

    if not session_id:
        return JsonResponse({'error': 'session_id is required.'}, status=400)

    session = _get_accessible_session(session_id, request.user)
    session.is_pinned = not session.is_pinned
    session.save()

    return JsonResponse({
        'system_action': 'chat_pinned',
        'session_id': session.id,
        'chat_metadata': {'is_pinned': session.is_pinned},
    })


# ─────────────────────────────────────────────
# Rename chat
# ─────────────────────────────────────────────

@login_required
@require_POST
def rename_chat_view(request):
    try:
        body = json.loads(request.body)
        session_id = body.get('session_id')
        new_title = body.get('new_title', '').strip()
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid request body.'}, status=400)

    if not session_id or not new_title:
        return JsonResponse({'error': 'session_id and new_title are required.'}, status=400)

    session = _get_accessible_session(session_id, request.user)
    session.title = new_title[:200]
    session.save()

    return JsonResponse({'system_action': 'chat_renamed', 'chat_title': session.title, 'session_id': session.id})


# ─────────────────────────────────────────────
# Delete current chat
# ─────────────────────────────────────────────

@login_required
@require_POST
def delete_current_chat_view(request):
    try:
        body = json.loads(request.body)
        session_id = body.get('session_id')
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid request body.'}, status=400)

    if not session_id:
        return JsonResponse({'error': 'session_id is required.'}, status=400)

    # Deleting is idempotent: if the chat is already gone (e.g. the button was
    # clicked twice, or another tab removed it), report success instead of a
    # 404 HTML page that the browser cannot parse as JSON.
    session = AnalysisSession.objects.filter(id=session_id).first()
    if session is None:
        return JsonResponse({'system_action': 'chat_deleted', 'session_id': session_id})

    # Only the chat's owner or a workspace member may delete it.
    is_owner = session.user == request.user
    is_member = session.workspace and WorkspaceMembership.objects.filter(
        workspace=session.workspace, user=request.user
    ).exists()
    if not (is_owner or is_member):
        return JsonResponse({'error': 'Access denied.'}, status=403)

    # Team (workspace) chats can only be deleted by the workspace owner.
    if session.workspace and session.workspace.owner != request.user:
        return JsonResponse(
            {'error': 'Only the workspace owner can delete team chats.'}, status=403
        )

    session.delete()

    return JsonResponse({'system_action': 'chat_deleted', 'session_id': session_id})


# ─────────────────────────────────────────────
# Mark a scenario done (tested) — saved & shared
# ─────────────────────────────────────────────

@login_required
@require_POST
def toggle_scenario_done_view(request):
    try:
        body = json.loads(request.body)
        db_id = body.get('db_id')
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid request.'}, status=400)

    scenario = get_object_or_404(TestScenario, id=db_id)
    session  = scenario.session

    if session.user != request.user:
        if not (session.workspace and WorkspaceMembership.objects.filter(
            workspace=session.workspace, user=request.user
        ).exists()):
            return JsonResponse({'error': 'Access denied.'}, status=403)

    scenario.is_done = not scenario.is_done
    scenario.save(update_fields=['is_done'])

    return JsonResponse({'status': 'ok', 'db_id': scenario.id, 'is_done': scenario.is_done})


# ─────────────────────────────────────────────
# Re-analyze an existing session
# ─────────────────────────────────────────────

@login_required
@require_POST
def reanalyze_view(request):
    try:
        body = json.loads(request.body)
        session_id = body.get('session_id')
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Invalid request.'}, status=400)

    session = _get_accessible_session(session_id, request.user)

    try:
        qa_result = QAAnalyzer().analyze(session.requirements_text)
    except GeminiError as e:
        return JsonResponse({'error': str(e)}, status=e.status)

    TrainingDataManager().reanalyze_session(session, qa_result)

    return JsonResponse(_build_session_response(
        session,
        requires_decision=session.accuracy_score < 80,
        suggested=session.suggested_requirement,
    ))


# ─────────────────────────────────────────────
# Export helpers — collect the full QA report once for all formats
# ─────────────────────────────────────────────

_ANALYSIS_FIELDS = [
    ('Actors', 'actors'),
    ('Actions', 'actions'),
    ('Business Rules', 'business_rules'),
    ('Constraints', 'constraints'),
    ('Validation Rules', 'validation_rules'),
    ('Error Handling', 'error_handling'),
    ('Non-Functional', 'non_functional'),
]


def _scenario_detailed_case(scenario):
    try:
        return scenario.detailed_case
    except DetailedTestCase.DoesNotExist:
        return None


def _collect_export_data(session):
    scenarios = list(session.test_scenarios.all())
    detailed = [(s.scenario_id, dc) for s in scenarios
                if (dc := _scenario_detailed_case(s)) is not None]
    requirements = _session_requirements(session)
    return {
        'title': session.title,
        'requirement_id': session.requirement_id or requirements[0].get('requirement_id', 'REQ-001'),
        'requirements_text': session.requirements_text,
        'generated_by': session.user.username,
        'created_at': session.created_at,
        'clarity': session.clarity_score,
        'completeness': session.completeness_score,
        'testability': session.testability_score,
        'overall': session.accuracy_score,
        'severity': session.severity or 'Medium',
        'label': session.get_score_label(),
        'requirements': requirements,
        'strengths': [f.message for f in session.feedback_items.filter(feedback_type='positive')],
        'warnings': [f.message for f in session.feedback_items.filter(feedback_type='warning')],
        'conditions': list(session.test_conditions.all()),
        'gaps': list(session.gaps.all()),
        'scenarios': scenarios,
        'detailed': detailed,
        'team_notes': session.team_notes,
    }


# ─────────────────────────────────────────────
# Export — PDF
# ─────────────────────────────────────────────

@login_required
def export_pdf_view(request, session_id):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    session = _get_accessible_session(session_id, request.user)
    d = _collect_export_data(session)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('ReportTitle', parent=styles['Title'], fontSize=16, spaceAfter=6)
    heading_style = ParagraphStyle('ReportHeading', parent=styles['Heading2'], fontSize=12, spaceBefore=8, spaceAfter=4)
    normal_style = ParagraphStyle('ReportNormal', parent=styles['Normal'], fontSize=9, leading=12)
    header_cell_style = ParagraphStyle('HeaderCell', parent=styles['Normal'], fontSize=8, leading=10,
                                       textColor=colors.white, fontName='Helvetica-Bold')
    body_cell_style = ParagraphStyle('BodyCell', parent=styles['Normal'], fontSize=7.5, leading=10,
                                      textColor=colors.black, wordWrap='CJK')

    table_style = TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f2f2f2')]),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
        ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 4),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
    ])

    def esc(t):
        return (str(t).replace('&', '&amp;').replace('<', '&lt;')
                .replace('>', '&gt;').replace('\n', '<br/>'))

    def make_table(headers, rows, col_widths):
        data = [[Paragraph(h, header_cell_style) for h in headers]]
        for row in rows:
            data.append([c if isinstance(c, Paragraph) else Paragraph(esc(c), body_cell_style) for c in row])
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(table_style)
        return t

    story = []

    # ── Header / summary ──
    story.append(Paragraph('QA Analysis &amp; Test Scenario Report', title_style))
    story.append(Paragraph(f'Project: {esc(d["title"])}', heading_style))
    story.append(Paragraph(
        f'Requirement ID: {esc(d["requirement_id"])} &nbsp;|&nbsp; Generated by: {esc(d["generated_by"])} '
        f'&nbsp;|&nbsp; {d["created_at"].strftime("%d %b %Y %H:%M")}', normal_style))
    story.append(Spacer(1, 0.3*cm))
    story.append(make_table(
        ['Clarity', 'Completeness', 'Testability', 'Overall', 'Severity'],
        [[f'{d["clarity"]}%', f'{d["completeness"]}%', f'{d["testability"]}%',
          f'{d["overall"]}% ({d["label"]})', d['severity']]],
        [3.4*cm]*5))
    story.append(Spacer(1, 0.4*cm))

    # ── 1. Requirement ──
    story.append(Paragraph('1. Requirement', heading_style))
    story.append(Paragraph(esc(d['requirements_text']), normal_style))

    # ── 2. Requirement Analysis (one block per requirement) ──
    if d['requirements']:
        story.append(Paragraph('2. Requirement Analysis', heading_style))
        for req in d['requirements']:
            rid = req.get('requirement_id', 'REQ-001')
            title = req.get('title', '')
            story.append(Paragraph('<b>' + esc(rid) + '</b>' + (' — ' + esc(title) if title else ''), normal_style))
            rows = [[label, '; '.join(str(x) for x in req.get(key, []))]
                    for label, key in _ANALYSIS_FIELDS if req.get(key)]
            if rows:
                story.append(make_table(['Field', 'Details'], rows, [4*cm, 13*cm]))
            story.append(Spacer(1, 0.2*cm))

    # ── 3. Quality Assessment ──
    if d['strengths'] or d['warnings']:
        story.append(Paragraph('3. Quality Assessment', heading_style))
        for s in d['strengths']:
            story.append(Paragraph(f'&#10003; {esc(s)}', normal_style))
        for wn in d['warnings']:
            story.append(Paragraph(f'&#9650; {esc(wn)}', normal_style))

    # ── 4. Test Conditions ──
    if d['conditions']:
        story.append(Paragraph('4. Test Conditions', heading_style))
        rows = [[c.condition_id, c.requirement_ref, c.description, c.condition_type, c.priority] for c in d['conditions']]
        story.append(make_table(['ID', 'Req', 'Condition', 'Type', 'Priority'], rows,
                                [1.4*cm, 1.8*cm, 8.4*cm, 2.6*cm, 2.4*cm]))

    # ── 5. Review Findings ──
    if d['gaps']:
        story.append(Paragraph('5. Requirement Review Findings', heading_style))
        rows = [[g.issue_id, g.issue_type, g.description, g.suggested_clarification] for g in d['gaps']]
        story.append(make_table(['ID', 'Type', 'Description', 'Suggested Clarification'], rows,
                                [1.4*cm, 2.8*cm, 6.4*cm, 6.4*cm]))

    # ── 6. Test Scenarios ──
    if d['scenarios']:
        story.append(Paragraph('6. Traceable Test Scenarios', heading_style))
        rows = []
        for s in d['scenarios']:
            meta = (f'<font size=6 color="#888888">Req: {esc(s.requirement_ref or "-")} &middot; '
                    f'Cond: {esc(s.condition_ref or "-")} &middot; {esc(s.scenario_type)} &middot; '
                    f'Done: {"yes" if s.is_done else "no"}</font><br/>')
            desc = Paragraph(meta + esc(s.description), body_cell_style)
            rows.append([s.scenario_id, desc, s.preconditions, s.expected_result, s.priority])
        story.append(make_table(
            ['ID', 'Description', 'Pre-Conditions', 'Expected Result', 'Priority'],
            rows, [1.4*cm, 6.0*cm, 4.3*cm, 3.8*cm, 1.5*cm]))

    # ── 7. Detailed Test Cases (only if generated) ──
    if d['detailed']:
        story.append(Paragraph('7. Detailed Test Cases', heading_style))
        rows = []
        for sid, dc in d['detailed']:
            steps_html = '<br/>'.join(f'{i+1}. {esc(step)}' for i, step in enumerate(dc.get_steps()))
            rows.append([sid, dc.test_data, steps_html, dc.expected_results, dc.postconditions])
        story.append(make_table(
            ['Scenario', 'Test Data', 'Steps', 'Expected Results', 'Postconditions'],
            rows, [1.8*cm, 3.4*cm, 4.4*cm, 3.7*cm, 3.7*cm]))

    # ── 8. Team Notes (only if any) ──
    if d['team_notes']:
        story.append(Paragraph('8. Team Notes', heading_style))
        story.append(Paragraph(esc(d['team_notes']), normal_style))

    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="qa_report_{session_id}.pdf"'
    return response


# ─────────────────────────────────────────────
# Export — Excel
# ─────────────────────────────────────────────

@login_required
def export_excel_view(request, session_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    session = _get_accessible_session(session_id, request.user)
    d = _collect_export_data(session)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
    label_font = Font(bold=True)
    top_wrap = Alignment(vertical='top', wrap_text=True)
    center_wrap = Alignment(horizontal='center', vertical='top', wrap_text=True)

    def header_row(ws, headers, row=1):
        for col, text in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_wrap

    def set_widths(ws, widths):
        for i, wdt in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = wdt

    def wrap_sheet(ws, start_row=2):
        for row in ws.iter_rows(min_row=start_row):
            for cell in row:
                cell.alignment = top_wrap

    wb = Workbook()

    # ── Sheet 1: Summary ──
    s1 = wb.active
    s1.title = 'Summary'
    summary = [
        ('Project', d['title']),
        ('Requirement ID', d['requirement_id']),
        ('Overall Score', f"{d['overall']}% ({d['label']})"),
        ('Clarity', f"{d['clarity']}%"),
        ('Completeness', f"{d['completeness']}%"),
        ('Testability', f"{d['testability']}%"),
        ('Severity', d['severity']),
        ('Generated By', d['generated_by']),
        ('Date', d['created_at'].strftime('%Y-%m-%d %H:%M')),
        ('Requirement', d['requirements_text']),
        ('Strengths', '\n'.join(d['strengths'])),
        ('Warnings', '\n'.join(d['warnings'])),
    ]
    for r, (label, value) in enumerate(summary, start=1):
        c1 = s1.cell(row=r, column=1, value=label); c1.font = label_font; c1.alignment = top_wrap
        s1.cell(row=r, column=2, value=value).alignment = top_wrap
    set_widths(s1, [18, 80])

    # ── Sheet 2: Requirement Analysis (one block per requirement) ──
    s2 = wb.create_sheet('Requirement Analysis')
    header_row(s2, ['Requirement', 'Field', 'Details'])
    r = 2
    for req in d['requirements']:
        rid = req.get('requirement_id', 'REQ-001')
        title = req.get('title', '')
        req_label = rid + (' — ' + title if title else '')
        for label, key in _ANALYSIS_FIELDS:
            s2.cell(row=r, column=1, value=req_label)
            s2.cell(row=r, column=2, value=label)
            s2.cell(row=r, column=3, value='\n'.join(str(x) for x in req.get(key, [])))
            r += 1
    wrap_sheet(s2)
    set_widths(s2, [24, 20, 80])

    # ── Sheet 3: Test Conditions ──
    s3 = wb.create_sheet('Test Conditions')
    header_row(s3, ['Condition ID', 'Requirement', 'Description', 'Type', 'Priority'])
    for r, c in enumerate(d['conditions'], start=2):
        s3.cell(row=r, column=1, value=c.condition_id)
        s3.cell(row=r, column=2, value=c.requirement_ref)
        s3.cell(row=r, column=3, value=c.description)
        s3.cell(row=r, column=4, value=c.condition_type)
        s3.cell(row=r, column=5, value=c.priority)
    wrap_sheet(s3)
    set_widths(s3, [14, 14, 62, 14, 12])

    # ── Sheet 4: Review Findings ──
    s4 = wb.create_sheet('Review Findings')
    header_row(s4, ['Issue ID', 'Issue Type', 'Description', 'Suggested Clarification'])
    for r, g in enumerate(d['gaps'], start=2):
        s4.cell(row=r, column=1, value=g.issue_id)
        s4.cell(row=r, column=2, value=g.issue_type)
        s4.cell(row=r, column=3, value=g.description)
        s4.cell(row=r, column=4, value=g.suggested_clarification)
    wrap_sheet(s4)
    set_widths(s4, [12, 18, 45, 45])

    # ── Sheet 5: Test Scenarios ──
    s5 = wb.create_sheet('Test Scenarios')
    header_row(s5, ['ID', 'Req Ref', 'Cond Ref', 'Description', 'Preconditions',
                    'Test Steps', 'Expected Result', 'Type', 'Priority', 'Done'])
    for r, s in enumerate(d['scenarios'], start=2):
        values = [s.scenario_id, s.requirement_ref, s.condition_ref, s.description,
                  s.preconditions, '\n'.join(s.get_steps()), s.expected_result,
                  s.scenario_type, s.priority, 'Yes' if s.is_done else 'No']
        for col, value in enumerate(values, start=1):
            s5.cell(row=r, column=col, value=value)
    wrap_sheet(s5)
    set_widths(s5, [10, 12, 10, 32, 28, 40, 32, 10, 10, 10])

    # ── Sheet 6: Detailed Cases (only if generated) ──
    if d['detailed']:
        s6 = wb.create_sheet('Detailed Cases')
        header_row(s6, ['Scenario ID', 'Test Data', 'Steps', 'Expected Results', 'Postconditions'])
        for r, (sid, dc) in enumerate(d['detailed'], start=2):
            s6.cell(row=r, column=1, value=sid)
            s6.cell(row=r, column=2, value=dc.test_data)
            s6.cell(row=r, column=3, value='\n'.join(dc.get_steps()))
            s6.cell(row=r, column=4, value=dc.expected_results)
            s6.cell(row=r, column=5, value=dc.postconditions)
        wrap_sheet(s6)
        set_widths(s6, [14, 35, 45, 35, 30])

    # ── Sheet 7: Team Notes (only if any) ──
    if d['team_notes']:
        s7 = wb.create_sheet('Team Notes')
        s7.cell(row=1, column=1, value=d['team_notes']).alignment = top_wrap
        set_widths(s7, [100])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="qa_report_{session_id}.xlsx"'
    return response
